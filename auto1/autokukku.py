#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Dec  1 15:26:27 2020

@author: shigematsuyuuki
"""


import openpyxl

wb=openpyxl.Workbook()
ws=wb['Sheet']

i=0
k=1
sum=0
for i in range(1,11) :
    t=ws.cell(row=i,column=2,value=k)
    k+=k
    if i==10 :
        break

for col in ws.iter_rows(min_col=2,max_row=10,max_col=2,values_only=True):
    for cell in col:
        print(cell)
        sum=sum+cell

ws['A11']='合計；'
ws['B11']=sum
wb.save('kuku.xlsx')
    
    