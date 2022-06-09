#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Nov 29 04:00:44 2020

@author: shigematsuyuuki
"""


import openpyxl
import docx

#ワークブックを追加する
wb=openpyxl.Workbook()

#シートを追加する
wb.create_sheet()

print(wb.sheetnames)

#アクティブシートを指定する
#sheet=wb.active
#辞書のキーを指定するようにシートを指定できる
sheet=wb['Sheet']


#セルの指定の仕方
#辞書的に指定
c=sheet['A1']

#cell()メソッドで1行1列を指定
#c=sheet.cell(raw=1,column=1)

#セルへの入力
c.value='Word'

sheet1=wb['Sheet1']
c1=sheet1['B1']
c1.value='Excel'

#保存
wb.save('Case1.xlsx')